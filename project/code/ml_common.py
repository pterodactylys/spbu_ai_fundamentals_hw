from __future__ import annotations

from dataclasses import dataclass
from math import sqrt
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from sklearn.compose import ColumnTransformer
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.impute import SimpleImputer
from sklearn.metrics import mean_absolute_error, mean_absolute_percentage_error, mean_squared_error
from sklearn.model_selection import train_test_split
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import OneHotEncoder, StandardScaler

from features import build_features

TABULAR_NUMERIC = [
    "ram_gb",
    "storage_gb",
    "iphone_gen",
    "photos_count",
    "title_len",
    "description_len",
    "days_since_posted",
    "is_pro",
    "is_max",
    "is_mini",
    "is_plus",
    "is_promoted",
    "has_warranty",
    "has_trade_in",
    "has_box",
    "has_charger",
]

TABULAR_CATEGORICAL = ["brand", "condition", "district"]

FEATURE_COLUMNS = ["title_only", "text"] + TABULAR_NUMERIC + TABULAR_CATEGORICAL


def load_dataset(path: Path) -> pd.DataFrame:
    if path.suffix.lower() in {".xlsx", ".xls"}:
        workbook = load_workbook(path, read_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        rows = list(worksheet.iter_rows(values_only=True))
        return pd.DataFrame(rows[1:], columns=rows[0])
    if path.suffix.lower() == ".csv":
        return pd.read_csv(path)
    raise ValueError(f"Unsupported format: {path}")


def prepare_dataset(path: Path) -> tuple[pd.DataFrame, pd.Series]:
    raw = load_dataset(path)
    data = build_features(raw)
    usable = data[FEATURE_COLUMNS + ["price"]].copy()
    X = usable.drop(columns=["price"])
    y = usable["price"]
    return X, y


def split_data(
    X: pd.DataFrame,
    y: pd.Series,
    *,
    test_size: float = 0.2,
    random_state: int = 42,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.Series, pd.Series]:
    return train_test_split(X, y, test_size=test_size, random_state=random_state)


def make_preprocessor(
    *,
    include_text: bool,
    text_column: str = "title_only",
    scale_numeric: bool = True,
    max_tfidf_features: int = 800,
) -> ColumnTransformer:
    numeric_steps: list = [("imputer", SimpleImputer(strategy="median"))]
    if scale_numeric:
        numeric_steps.append(("scaler", StandardScaler()))

    transformers: list = [
        ("num", Pipeline(numeric_steps), TABULAR_NUMERIC),
        (
            "cat",
            Pipeline([
                ("imputer", SimpleImputer(strategy="most_frequent")),
                ("onehot", OneHotEncoder(handle_unknown="ignore", max_categories=30)),
            ]),
            TABULAR_CATEGORICAL,
        ),
    ]

    if include_text:
        transformers.append(
            (
                "text",
                TfidfVectorizer(
                    max_features=max_tfidf_features,
                    ngram_range=(1, 2),
                    min_df=2,
                    strip_accents="unicode",
                ),
                text_column,
            )
        )

    return ColumnTransformer(transformers)


def metrics_in_rubles(y_true: np.ndarray, y_pred: np.ndarray) -> dict:
    y_true = np.asarray(y_true, dtype=float)
    y_pred = np.clip(np.asarray(y_pred, dtype=float), 0, None)
    return {
        "mape": float(mean_absolute_percentage_error(y_true, y_pred)),
        "mae": float(mean_absolute_error(y_true, y_pred)),
        "rmse": float(sqrt(mean_squared_error(y_true, y_pred))),
    }


@dataclass
class BoostingFeatures:

    vectorizer: TfidfVectorizer
    numeric_imputer: SimpleImputer
    tfidf_columns: list[str]

    def transform(self, X: pd.DataFrame) -> pd.DataFrame:
        numeric = pd.DataFrame(
            self.numeric_imputer.transform(X[TABULAR_NUMERIC]),
            columns=TABULAR_NUMERIC,
            index=X.index,
        )
        categorical = X[TABULAR_CATEGORICAL].fillna("unknown").astype(str)
        for col in TABULAR_CATEGORICAL:
            categorical[col] = categorical[col].astype("category")
        tfidf = pd.DataFrame(
            self.vectorizer.transform(X["title_only"]).toarray(),
            columns=self.tfidf_columns,
            index=X.index,
        )
        return pd.concat([numeric, categorical, tfidf], axis=1)


def fit_boosting_features(
    X_train: pd.DataFrame,
    *,
    max_tfidf_features: int = 300,
) -> BoostingFeatures:
    imputer = SimpleImputer(strategy="median")
    imputer.fit(X_train[TABULAR_NUMERIC])

    vectorizer = TfidfVectorizer(
        max_features=max_tfidf_features,
        ngram_range=(1, 2),
        min_df=2,
        strip_accents="unicode",
    )
    vectorizer.fit(X_train["title_only"])
    tfidf_columns = [f"tf_{i}" for i in range(len(vectorizer.get_feature_names_out()))]

    return BoostingFeatures(
        vectorizer=vectorizer,
        numeric_imputer=imputer,
        tfidf_columns=tfidf_columns,
    )


def evaluate_boosting(
    name: str,
    model,
    features: BoostingFeatures,
    X_train: pd.DataFrame,
    X_test: pd.DataFrame,
    y_train: pd.Series,
    y_test: pd.Series,
    *,
    use_log_target: bool,
    fit_kwargs: dict | None = None,
) -> dict:
    fit_kwargs = fit_kwargs or {}
    X_tr = features.transform(X_train)
    X_te = features.transform(X_test)
    y_fit = np.log1p(y_train.values) if use_log_target else y_train.values

    model.fit(X_tr, y_fit, **fit_kwargs)
    preds = model.predict(X_te)

    if use_log_target:
        preds_eval = np.expm1(np.clip(preds, 0, 15))
    else:
        preds_eval = np.clip(preds, 0, None)

    return {
        "model": name,
        **metrics_in_rubles(y_test.values, preds_eval),
        "predictions": preds_eval,
        "fitted_model": model,
        "features": features,
    }


def evaluate_model(
    name: str,
    model: Pipeline,
    X_train: pd.DataFrame,
    X_test: pd.DataFrame,
    y_train: pd.Series,
    y_test: pd.Series,
    *,
    use_log_target: bool,
) -> dict:
    y_fit = np.log1p(y_train.values) if use_log_target else y_train.values
    model.fit(X_train, y_fit)
    preds = model.predict(X_test)

    if use_log_target:
        y_test_eval = y_test.values
        preds_eval = np.expm1(np.clip(preds, 0, 15))
    else:
        y_test_eval = y_test.values
        preds_eval = np.clip(preds, 0, None)

    return {
        "model": name,
        **metrics_in_rubles(y_test_eval, preds_eval),
        "predictions": preds_eval,
        "fitted_model": model,
    }
